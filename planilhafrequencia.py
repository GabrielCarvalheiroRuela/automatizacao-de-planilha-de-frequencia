#C:\Users\gabri\AppData\Local\Packages\PythonSoftwareFoundation.Python.3.11_qbz5n2kfra8p0\LocalCache\local-packages\Python311\Scripts\pyinstaller.exe --onefile --noconsole .\planilhafrequencia.py
import datetime
import calendar
import time
from tkinter import *
import tkinter as tk
from tkinter import ttk
import openpyxl
from openpyxl import Workbook
from openpyxl.drawing.image import Image
import threading
import os
import shutil

def create_spreadsheet():
    message_label.config(text="Gerando planilha ... aguarde!")
    name = entry_name.get()
    sector = entry_sector.get()
    period = workperiod.get()
    day1 = day1notwork.get()
    day2 = day2notwork.get()
    day3 = day3notwork.get()
    day4 = day4notwork.get()
    day5 = day5notwork.get()
    cause1 = cause1notworking.get()
    cause2 = cause2notworking.get()
    cause3 = cause3notworking.get()
    cause4 = cause4notworking.get()
    cause5 = cause5notworking.get()

    # Definindo mês atual e passado
    current_date = datetime.datetime.now()
    numb_current_month = current_date.month
    numb_last_month = current_date.month - 1
    if numb_current_month == 1:
        current_month = 'Janeiro'
        last_month = 'Dezembro'
    elif numb_current_month == 2:
        current_month = 'Fevereiro'
        last_month = 'Janeiro'
    elif numb_current_month == 3:
        current_month = 'Março'
        last_month = 'Fevereiro'
    elif numb_current_month == 4:
        current_month = 'Abril'
        last_month = 'Março'
    elif numb_current_month == 5:
        current_month = 'Maio'
        last_month = 'Abril'
    elif numb_current_month == 6:
        current_month = 'Junho'
        last_month = 'Maio'
    elif numb_current_month == 7:
        current_month = 'Julho'
        last_month = 'Junho'
    elif numb_current_month == 8:
        current_month = 'Agosto'
        last_month = 'Julho'
    elif numb_current_month == 9:
        current_month = 'Setembro'
        last_month = 'Agosto'
    elif numb_current_month == 10:
        current_month = 'Outubro'
        last_month = 'Setembro'
    elif numb_current_month == 11:
        current_month = 'Novembro'
        last_month = 'Outrubro'
    elif numb_current_month == 12:
        current_month = 'Dezembro'
        last_month = 'Novembro'

    # Definindo quantidade de dias que o mês passado possui
    current_year = current_date.year
    if numb_current_month == 1:
        numb_last_month = 12
        last_year = current_year - 1
    else:
        numb_last_month = numb_current_month - 1
        last_year = current_year
    
    start_day = 21
    days_last_month = calendar.monthrange(last_year, numb_last_month)[1]
    week_day21 = calendar.weekday(last_year, numb_current_month, start_day)

    days_of_week = ["Segunda-feira", "Terça-feira", "Quarta-feira", "Quinta-feira", "Sexta-feira", "Sábado", "Domingo"]

    # Definindo periodo de trabalho dos estagiários
    if period == 'morning':
        entry_time = '7:00'
        departure_time = '13:00'
    if period == 'afternoon':
        entry_time = '13:00'
        departure_time = '19:00'
    if period == 'night':
        entry_time = '16:00'
        departure_time = '22:00'
    
    workbook = openpyxl.load_workbook('Template.xlsx')
    sheet = workbook.active
    # -------------------------- ADICIONAR DADOS NA TABELA ----------------------------
    # Dias do mes
    for linha in range(14,45):
        if linha < 25:
            cells_day_value = sheet.cell(row=linha, column=2)
            cells_day_value.value = linha + 7
        else:
            cells_day_value = sheet.cell(row=linha, column=2)
            cells_day_value.value = linha - 24

    # LOGO UNAERP
    img = Image('logo.jpeg')
    img.anchor = 'A1'
    sheet.add_image(img)

    # Dados principais do estagiário
    sheet['C6'] = name
    sheet['B5'] = f"{last_month}/{current_month}"
    sheet['E5'] = current_date.year
    sheet['D7'] = sector

    # Dados dos meses
    for linha in range(14, 25):
        if sheet[f"B{linha}"].value <= days_last_month:
            cells_months = sheet.cell(row=linha, column=1)
            cells_months.value = last_month

    for linha in range(25, 45):
        cells_months = sheet.cell(row=linha, column=1)
        cells_months.value = current_month

    # Dados das semanas
    start_day = 21
    for i in range(11):
        if start_day <= days_last_month:
            week_day21 = calendar.weekday(last_year, numb_last_month, start_day)
            name_week_day = days_of_week[week_day21]
            cells_week = sheet.cell(row=i+14, column=3)
            cells_week.value = name_week_day
            start_day += 1

    start_day = 1
    for i in range(20):
        if start_day <= calendar.monthrange(current_year, numb_current_month)[1]:
            week_day21 = calendar.weekday(current_year, numb_current_month, start_day)
            name_week_day = days_of_week[week_day21]
            cells_week = sheet.cell(row=i+25, column=3)
            cells_week.value = name_week_day
            start_day += 1

    # Dados de entrada e saída do estagiário
    for linha in range(14, 45):
        if (sheet[f"C{linha}"].value == 'Domingo') or (sheet[f"C{linha}"].value == 'Sábado') or (sheet[f"C{linha}"].value == None):
            cells_entry_time = sheet.cell(row=linha, column=4)
            cells_departure_time = sheet.cell(row=linha, column=5)
            cells_entry_time.value = ""
            cells_departure_time.value = ""
        else:
            cells_entry_time = sheet.cell(row=linha, column=4)
            cells_departure_time = sheet.cell(row=linha, column=5)
            cells_entry_time.value = entry_time
            cells_departure_time.value = departure_time 

    # Dias que não foram trabalhados
    for linha in range(14, 45):
        if str(sheet[f"B{linha}"].value) == day1:
            sheet[f"D{linha}"].value = cause1
            sheet[f"E{linha}"].value = cause1
        if str(sheet[f"B{linha}"].value) == day2:
            sheet[f"D{linha}"].value = cause2
            sheet[f"E{linha}"].value = cause2
        if str(sheet[f"B{linha}"].value) == day3:
            sheet[f"D{linha}"].value = cause3
            sheet[f"E{linha}"].value = cause3
        if str(sheet[f"B{linha}"].value) == day4:
            sheet[f"D{linha}"].value = cause4
            sheet[f"E{linha}"].value = cause4
        if str(sheet[f"B{linha}"].value) == day5:
            sheet[f"D{linha}"].value = cause5
            sheet[f"E{linha}"].value = cause5

    workbook.save(f"{name} - {current_month} - Planilha de frequência.xlsx")
    downloads_path = os.path.expanduser("~\\Downloads")
    caminho_origem = f"{name} - {current_month} - Planilha de frequência.xlsx"
    novo_caminho = os.path.join(downloads_path, caminho_origem)
    shutil.move(caminho_origem, novo_caminho)
    workbook.close()
    # Aviso de planilha criada com sucesso
    message_label.config(text="Planilha gerada com sucesso!", fg="green")
    x=5
    for i in range(5):
        message_close_label.config(text=f"Fechando em {x} segundos", fg="green")
        time.sleep(1)
        x -= 1
    windows.destroy()

def generate_spreadsheet():
    # Desativa o botão enquanto a planilha está sendo gerada
    button.config(state=DISABLED)
    create_spreadsheet()
    # Reativa o botão após a planilha ser gerada
    button.config(state=NORMAL)

def on_generate_button_click():
    t = threading.Thread(target=generate_spreadsheet)
    t.start()

windows = Tk()
windows.title("Planilha de frequência")

# Nome do usuário
label_name = Label(windows, text="Nome completo:")
label_name.grid(row=0, column=0, padx=10, pady=(10,5), sticky='w')
entry_name = Entry(windows, width=30)
entry_name.grid(row=0, column=1, padx=10, pady=(10,5), sticky='w')

# Setor que o usuário trabalha
label_sector = Label(windows, text="Setor:")
label_sector.grid(row=1, column=0, padx=10, pady=5, sticky='w')
entry_sector = Entry(windows, width=30)
entry_sector.grid(row=1, column=1, columnspan=2, padx=10, pady=5, sticky='w')

# Período que o usuário trabalha
workperiod = StringVar()
workperiod.set("")
label_sector = Label(windows, text="Período:")
label_sector.grid(row=2, column=0, padx=10, pady=5, sticky='w')
radio_morning = ttk.Radiobutton(windows, text="Manhã", variable=workperiod, value="morning")
radio_morning.grid(row=2, column=1, padx=10, pady=5, sticky='w')
radio_afternoon = ttk.Radiobutton(windows, text="Tarde", variable=workperiod, value="afternoon")
radio_afternoon.grid(row=2, column=1, padx=10, pady=5, sticky='n')
radio_night = ttk.Radiobutton(windows, text="Noite", variable=workperiod, value="night")
radio_night.grid(row=2, column=1, padx=5, pady=5, sticky='e')

# Férias, Recessos, Atestados
notwork = Label(windows, text="(FERIADO, RECESSO, ATESTADO, FALTA)", font=("Arial", 10, "bold"))
notwork.grid(row=3, column=0, columnspan=2, padx=10)
inf_day = Label(windows, text="Dia  ", font=("Arial", 10, "bold"))
inf_day.grid(row=4, column=0, padx=10, sticky='e')
inf_cause = Label(windows, text="        Causa", font=("Arial", 10, "bold"))
inf_cause.grid(row=4, column=1, padx=10, sticky='w')

day1notwork = Entry(windows, width=10)
day1notwork.grid(row=5,column=0,sticky="e")
cause1notworking = Entry(windows, width=20)
cause1notworking.grid(row=5, column=1, padx=5,sticky="w")

day2notwork = Entry(windows, width=10)
day2notwork.grid(row=6,column=0,sticky="e")
cause2notworking = Entry(windows, width=20)
cause2notworking.grid(row=6, column=1, padx=5,sticky="w")

day3notwork = Entry(windows, width=10)
day3notwork.grid(row=7,column=0,sticky="e")
cause3notworking = Entry(windows, width=20)
cause3notworking.grid(row=7, column=1, padx=5,sticky="w")

day4notwork = Entry(windows, width=10)
day4notwork.grid(row=8,column=0,sticky="e")
cause4notworking = Entry(windows, width=20)
cause4notworking.grid(row=8, column=1, padx=5,sticky="w")

day5notwork = Entry(windows, width=10)
day5notwork.grid(row=9,column=0,sticky="e")
cause5notworking = Entry(windows, width=20)
cause5notworking.grid(row=9, column=1, padx=5,sticky="w")

# Botão para gerar planilha
button = Button(windows, text="Gerar planilha", command=on_generate_button_click, font=("Arial", 10, "bold"))
button.grid(row=10, column=0, columnspan=2, pady=(10, 0))

# Créditos
credits = Label(windows, text="Credits - by Gabriel Ruela", font=("Arial", 8, "bold"))
credits.grid(row=11, column=0, columnspan=2, pady=(0))

# Aviso de planilha gerada
message_label = Label(windows, text="", font=("Arial", 10, "bold"))
message_label.grid(row=12, column=0, columnspan=2, pady=(5,0))
message_close_label = Label(windows, text="", font=("Arial", 7, "bold"))
message_close_label.grid(row=13, column=0, columnspan=2, pady=(0,5))

windows.mainloop()